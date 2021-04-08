import openpyxl
from urllib.parse import urlparse
wb = openpyxl.load_workbook("50000.xlsx")
ws = wb.active
rows = ws.iter_rows()
wb2 = openpyxl.Workbook()
ws2 = wb2.active
title_list = next(rows)
t_l = list(map(lambda x:x.value,title_list))
t_l.append("ID")
ws2.append(t_l)
for row in rows:
    if row[0] is None:
        continue
    i = row[11].value
    l = list(map(lambda x:x.value,row))
    p = urlparse(i).path
    u = p.strip("/")

    id = u.split("/")[-1]
    print(id)
    try:
        id = int(id)
    except Exception:
        ws2.append(l)
        continue
    id = str(id)
    l.append(id)
    ws2.append(l)

wb2.save("50000-2.xlsx")