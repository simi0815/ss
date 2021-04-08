import openpyxl
from searchsystem.api.utils.lib.all_sql import AllDatabase

class XML_POST_HANDER(object):
    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file)
        self.sheet = self.wb.active
    def _get_name_index(self,title_row):
        print(title_row)
        name_index_dict = {
            "ch_name_index":-1,
            "first_name":-1,
            "last_name":-1
        }
        for i in range(len(title_row)):
            if title_row[i].value == '中文姓名':
                name_index_dict["ch_name_index"] = i
            elif title_row[i].value == '英文姓':
                name_index_dict["first_name"] = i
            elif title_row[i].value == '英文名':
                name_index_dict["last_name"] = i
        return name_index_dict
    def _get_name(self,n_i_dict,row):

        if n_i_dict["first_name"] != -1:
            f_name = row[n_i_dict["first_name"]].value
            l_name = row[n_i_dict["last_name"]].value
            if f_name :
                en_name =" ".join([f_name,l_name])
                en_name = en_name.strip(' ')
                if en_name:
                    return en_name
        if n_i_dict["ch_name_index"] != -1:
            ch_name = row[n_i_dict["ch_name_index"]].value
            if ch_name:
                ch_name = ch_name.strip(' ')
                return ch_name
        return ""

    def get_res(self):
        rows = self.sheet.iter_rows()
        n_i_dict = self._get_name_index(next(rows))
        for row in rows:
            name = self._get_name(n_i_dict,row)
            adb = AllDatabase()
            if name and isinstance(name,str):
                res = adb.find({"val":name,"type":"name"})
                if res["errno"] == 0:
                    count = res["data"]["count"]
                    print(name,count)
                else:
                    print(res["msg"])







if __name__ == '__main__':
    xph = XML_POST_HANDER("3000.xlsx")
    xph.get_res()

# wb = openpyxl.load_workbook('3000.xlsx')
# sheet = wb.active
# rows = sheet.iter_rows()
# wb2 = openpyxl.Workbook()
# ws = wb2.active
#
# title_row = next(rows)
# row_num = 0
# for row in rows:
#     row_num = row_num + 1
#     name = ''
#     if row[3].value and row[4].value:
#         name = row[3].value + ' ' + row[4].value
#     if not name:
#         name = row[2].value
#     if name:
#         res = requests.get("http://192.168.2.88:8000/api/v1/infolist", params={
#             "val": name,
#             "type": "name"
#         })
#
#         j_res = res.json()
#
#         if (j_res.get("errno") == 0):
#             count = len(j_res["data"]["all_data"])
#             print(name, count)
#             ws.append([name, count])
# wb2.save("res4.xlsx")
