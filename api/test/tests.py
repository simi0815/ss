import openpyxl
from openpyxl.styles.fonts import Font
from openpyxl.styles import colors
ws = openpyxl.load_workbook("text.xlsx")
wb = ws.active

rows = wb.iter_rows()
print(wb.max_row)
for row in rows:
    print(row)


