import openpyxl
from api.lib.all_sql import AllDatabase
from api.lib.data.toxml import StoreToXML
from api.utils.xml import store_xml
from api.lib.redis_db import RedisDB
import os
FIELD_LIST = ["姓名", "地址", "出生日期", "性别", "电话", "邮箱", "用户名", "社交ID", "社交类型", "数据来源"]
FIELD_STREBTH_INDEX = {
    "name": 1,
    "tel": 5,
    "email": 6,
    "uname": 7,
    "cid": 8
}


class XmlPostHander(object):
    def __init__(self, file, select_field):
        self.file_name = os.path.basename(file)
        self.wb = openpyxl.load_workbook(file)
        self.sheet = self.wb.active
        self.select_field = select_field
        self.red = RedisDB()

    def get_res(self):
        rows = self.sheet.iter_rows()
        self.red.set_one(self.file_name,"all",self.sheet.max_row-1)
        self.red.set_one(self.file_name, "status", "正在查询中...")
        al_data = []
        count = 0
        err_count = 0
        # 去掉标题行
        next(rows)
        for row in rows:
            count = count + 1
            print("当前查询条数：", count)
            self.red.set_one(self.file_name,"cur",count)
            try:
                adb = AllDatabase()
                for field, col_info in self.select_field.items():
                    if col_info.get("isChecked") and col_info.get("col") + 1:
                        col = col_info["col"]
                    else:
                        continue
                    ctype = col_info.get("ctype")
                    col_val = str(row[col].value)
                    if col_val:
                        res = adb.find({"val": col_val, "type": field, "ctype": ctype})
                        if res["errno"] == 0:
                            stx = StoreToXML(res["data"]["all_data"])
                            data_list = stx.paser_content()
                            al_data.extend(data_list)
                self.red.set_one(self.file_name, "res_count", len(al_data))
                print("共得到{count}条结果".format(count=len(al_data)))
            except Exception as e:
                err_count += 1
                msg = "查询第{count}条时出现错误:{e}".format(count=count, e=e)
                self.red.set_one(self.file_name, "err_count", err_count)
                self.red.set_one(self.file_name, "cur_msg", msg)
                print(msg)
                al_data.append([msg])
        self.red.set_one(self.file_name, "status", "查询完毕，正在生成结果文件...")
        url = store_xml(al_data, FIELD_LIST, "ALL_INFO")
        self.red.set_one(self.file_name, "status", "查询结束")
        self.red.set_one(self.file_name, "url", url)
        return url
