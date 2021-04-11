from openpyxl import Workbook
from openpyxl.styles.fonts import Font
import os
import uuid
from api.conf.api_config import STATIC_ROOT, PROTOCOL_IP_OR_DOMAIN, STATIC_URL
from api.utils.date import get_cur_date


def store_xml(data, first_row, sheet_name, strenth_col=None, strenth_color='#ff0000'):
    wb = Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    ws.append(first_row)
    for onr in data:
        ws.append(onr)
    # 关键标红
    if strenth_col:
        col = chr(64+strenth_col)
        if col < 'A' or col > 'Z':
            raise Exception("不支持标红列数多于26")
        colB = ws[col]
        for cell in colB:
            # cell.value = 'sb'
            cell.font = Font(color="ff0000")
    name = uuid.uuid4()
    file_name = str(name) + ".xlsx"
    date_dir = get_cur_date()
    file_dir = os.path.join(STATIC_ROOT, date_dir)
    if not os.path.exists(file_dir):
        os.makedirs(file_dir, exist_ok=True)
    path = os.path.join(file_dir, file_name)
    wb.save(path)
    url = PROTOCOL_IP_OR_DOMAIN + STATIC_URL + "/" + date_dir + "/" + file_name
    return url
